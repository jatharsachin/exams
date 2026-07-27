"""
SPPU Exam Management System - Utility Functions
"""

from datetime import datetime, date
import os


def format_date(d):
    if isinstance(d, str):
        return d
    if isinstance(d, (date, datetime)):
        return d.strftime("%d-%m-%Y")
    return str(d)


def today_str():
    return datetime.now().strftime("%d-%m-%Y")


def export_to_excel(data, columns, filepath, sheet_name="Data"):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for ci, col in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col)
    for ri, row in enumerate(data, 2):
        for ci, key in enumerate(columns, 1):
            ws.cell(row=ri, column=ci, value=row.get(key, ""))
    wb.save(filepath)


def alphabet_range(n):
    result = []
    while n >= 0:
        result.append(chr(65 + (n % 26)))
        n = n // 26 - 1
    return "".join(reversed(result))
